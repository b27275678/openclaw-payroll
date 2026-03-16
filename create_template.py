import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Payroll"

headers = ["Employee Name", "Employee ID", "Hourly Rate", "Regular Hours", "Overtime Hours", "Holiday Hours", "BonusTravel Hours", "BonusTravel Rate", "ToolsPurch", "LevyGarn", "Insurance"]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions[cell.column_letter].width = 18

ws.append(["BRANDON SINNER", "BRANDONS-XXX-XX-6781", 20.00, 75.25, 11.25, 8.00, 0, 0, 50.00, 675.85, 16.00])

wb.save("payroll_template.xlsx")
print("Template created!")
